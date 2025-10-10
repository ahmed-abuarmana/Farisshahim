from django.shortcuts import render, get_object_or_404, redirect
from .models import Makhbaz, Takiya, Taslima_makhbaz, Taslima_takiya
from django.http import JsonResponse
import json
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging
import urllib.parse



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('index')  # يرجع للصفحة الرئيسية بعد الدخول
        else:
            return render(request, 'login.html', {'error': '❌ اسم المستخدم أو كلمة المرور غير صحيحة'})
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')  # يرجع لصفحة تسجيل الدخول



@login_required(login_url='login')
def index(request):
    makhabiz = Makhbaz.objects.all()
    takiyat = Takiya.objects.all()
    context = {
        "makhabiz": makhabiz,
        "takiyat": takiyat,
    }
    return render(request, "index.html", context)



@login_required(login_url='login')
# عرض كل المخابز
def makhabez_list(request):
    makhabiz = Makhbaz.objects.all()
    return render(request, "makhabez_list.html", {"makhabiz": makhabiz})



@login_required(login_url='login')
# عرض تفاصيل مخبز محدد مع جميع التسليمات
def makhabez_detail(request, pk):
    makhbaz = get_object_or_404(Makhbaz, pk=pk)

    # جلب التسليمات المرتبطة بهذا المخبز
    all_tasleemat = makhbaz.taslimat.all().order_by('-taslima_date')

    latest_taslim = all_tasleemat.first() if all_tasleemat.exists() else None

    # حساب الإجماليات
    total_flour = sum(t.flour or 0 for t in all_tasleemat)
    total_salt = sum(t.salt or 0 for t in all_tasleemat)
    total_yeast = sum(t.yeast or 0 for t in all_tasleemat)
    total_sugar = sum(t.sugar or 0 for t in all_tasleemat)
    total_cooking_oil = sum(t.cooking_oil or 0 for t in all_tasleemat)
    total_wood = sum(t.wood or 0 for t in all_tasleemat)
    total_gas = sum(t.gas or 0 for t in all_tasleemat)

    context = {
        "makhbaz": makhbaz,
        "all_tasleemat": all_tasleemat,
        "latest_taslim": latest_taslim,
        "total_deliveries": all_tasleemat.count(),
        "total_flour": total_flour,
        "total_salt": total_salt,
        "total_yeast": total_yeast,
        "total_sugar": total_sugar,
        "total_cooking_oil": total_cooking_oil,
        "total_wood": total_wood,
        "total_gas": total_gas,
    }

    return render(request, "makhabez_detail.html", context)



@login_required(login_url='login')
# @require_http_methods(["POST"])
def update_makhbaz(request, pk):
    if request.method == "POST":
        try:
            makhbaz = get_object_or_404(Makhbaz, pk=pk)
            data = json.loads(request.body)
            
            # تحديث الحقول
            fields = [
                'name', 'owner_name', 'owner_id', 'mobile_number', 
                'address', 'governorate', 'coordinates', 'oven_type',
                'production_capacity', 'contract_type', 'status'
            ]
            
            for field in fields:
                if field in data:
                    setattr(makhbaz, field, data[field])
            
            makhbaz.save()
            
            return JsonResponse({"success": True})
            
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})



@login_required(login_url='login')
@require_http_methods(["POST"])
def update_takiya(request, pk):
    """
    تحديث بيانات التكية المحددة بواسطة مفتاحها الأساسي (pk).
    تستقبل البيانات المعدلة كـ JSON في جسم الطلب.
    """
    try:
        takiya = get_object_or_404(Takiya, pk=pk)
        data = json.loads(request.body)
        
        # قائمة بالحقول القابلة للتعديل
        fields = [
            'name', 'owner_name', 'owner_id', 'mobile_number', 
            'address', 'governorate', 'coordinates', 'status',
            'total_pots', 'pots_80', 'pots_100', 'pots_120', 
            'pots_150', 'pots_200', 'daily_capacity'
        ]
        
        for field in fields:
            # التحقق مما إذا كان الحقل موجودًا في بيانات JSON
            # واستخدام قيمة None إذا كان الحقل فارغًا أو غير موجود لتجنب أخطاء التحويل
            if field in data:
                value = data[field] if data[field] != "" else None
                setattr(takiya, field, value)
        
        # ملاحظة: حقول الاختيار (choices) مثل 'governorate' و 'status' يجب أن تكون 
        # قيمتها في JSON هي القيمة المخزنة (مثل "رفح" أو "فعال").

        takiya.save()
        
        return JsonResponse({"success": True})
            
    except Takiya.DoesNotExist:
        return JsonResponse({"success": False, "error": "التكية غير موجودة"}, status=404)
    except Exception as e:
        # يمكن تسجيل الخطأ التفصيلي في السجل (logging)
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required(login_url='login')
# عرض كل التكيات
def takiyat_list(request):
    takiyat = Takiya.objects.all()
    return render(request, "takiyat_list.html", {"takiyat": takiyat})



@login_required(login_url='login')
# عرض تفاصيل تكية محددة مع جميع التسليمات
def takiya_detail(request, pk):
    takiya = get_object_or_404(Takiya, pk=pk)

    # جلب التسليمات المرتبطة بهذه التكية (من Taslima_takiya)
    all_tasleemat = takiya.taslimat.all().order_by('-taslima_date')

    latest_taslim = all_tasleemat.first() if all_tasleemat.exists() else None

    # حساب الإجماليات لجميع الحقول
    total_salt = sum(t.salt or 0 for t in all_tasleemat)
    total_macaroni = sum(t.macaroni or 0 for t in all_tasleemat)
    total_rice = sum(t.rice or 0 for t in all_tasleemat)
    total_oil = sum(t.oil or 0 for t in all_tasleemat)
    total_peas = sum(t.peas or 0 for t in all_tasleemat)
    total_lentils = sum(t.lentils or 0 for t in all_tasleemat)
    total_beans = sum(t.beans or 0 for t in all_tasleemat)
    total_sauce = sum(t.sauce or 0 for t in all_tasleemat)
    total_luncheon = sum(t.luncheon or 0 for t in all_tasleemat)
    total_maggi_spice = sum(t.maggi_spice or 0 for t in all_tasleemat)
    total_vegetable_soup = sum(t.vegetable_soup or 0 for t in all_tasleemat)
    total_seven_spices = sum(t.seven_spices or 0 for t in all_tasleemat)
    total_ghee = sum(t.ghee or 0 for t in all_tasleemat)
    total_bulgur = sum(t.bulgur or 0 for t in all_tasleemat)

    context = {
        "takiya": takiya,
        "all_tasleemat": all_tasleemat,
        "latest_taslim": latest_taslim,
        "total_deliveries": all_tasleemat.count(),
        
        # إجماليات التسليمات
        "total_salt": total_salt,
        "total_macaroni": total_macaroni,
        "total_rice": total_rice,
        "total_oil": total_oil,
        "total_peas": total_peas,
        "total_lentils": total_lentils,
        "total_beans": total_beans,
        "total_sauce": total_sauce,
        "total_luncheon": total_luncheon,
        "total_maggi_spice": total_maggi_spice,
        "total_vegetable_soup": total_vegetable_soup,
        "total_seven_spices": total_seven_spices,
        "total_ghee": total_ghee,
        "total_bulgur": total_bulgur,
        
        # بيانات التكية الأساسية
        "governorate_choices": Takiya.GOVERNORATE_CHOICES,
        "status_choices": Takiya.STATUS_CHOICES,
        
        # إجماليات القدور
        "total_pots_all": (takiya.pots_80 or 0) + (takiya.pots_100 or 0) + 
                         (takiya.pots_120 or 0) + (takiya.pots_150 or 0) + 
                         (takiya.pots_200 or 0),
    }

    return render(request, "takiya_detail.html", context)


@login_required(login_url='login')
def add_tasleema_for_makhbaz(request, makhbaz_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            makhbaz = Makhbaz.objects.get(id=makhbaz_id)
            
            # إنشاء التسليمة
            taslima = Taslima_makhbaz.objects.create(
                taslima_date = data.get('taslima_date'),
                flour = data.get('flour') or None,
                yeast = data.get('yeast') or None,
                salt = data.get('salt') or None,
                sugar = data.get('sugar') or None,
                cooking_oil = data.get('cooking_oil') or None,
                wood = data.get('wood') or None,
                gas = data.get('gas') or None,
                additions = data.get('additions') or None,
                makhbaz = makhbaz
            )
            
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


@login_required(login_url='login')
def add_tasleema_for_takiya(request, takiya_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            takiya = Takiya.objects.get(id=takiya_id)
            
            taslima = Taslima_takiya.objects.create(
                taslima_date=data.get('taslima_date'),
                salt=data.get('salt') or None,
                macaroni=data.get('macaroni') or None,
                rice=data.get('rice') or None,
                oil=data.get('oil') or None,
                peas=data.get('peas') or None,
                lentils=data.get('lentils') or None,
                beans=data.get('beans') or None,
                sauce=data.get('sauce') or None,
                luncheon=data.get('luncheon') or None,
                maggi_spice=data.get('maggi_spice') or None,
                vegetable_soup=data.get('vegetable_soup') or None,
                seven_spices=data.get('seven_spices') or None,
                ghee=data.get('ghee') or None,
                bulgur=data.get('bulgur') or None,
                additions=data.get('additions') or None,  # إضافة حقل الإضافات
                takiya=takiya
            )
            return JsonResponse({"success": True, "taslima_id": taslima.id})
        except Takiya.DoesNotExist:
            return JsonResponse({"success": False, "error": "التكية غير موجودة"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "طريقة الطلب غير صالحة"})


@login_required(login_url='login')
def add_new_makhbaz(request):
    if request.method == "POST":
        data = {
            'name': request.POST.get("name") or None,
            'governorate': request.POST.get("governorate") or None,
            'address': request.POST.get("address") or None,
            'owner_name': request.POST.get("owner_name") or None,
            'owner_id': request.POST.get("owner_id") or None,
            'mobile_number': request.POST.get("mobile_number") or None,
            'coordinates': request.POST.get("coordinates") or None,
            'oven_type': request.POST.get("oven_type") or None,
            'production_capacity': request.POST.get("production_capacity") or None,
            'contract_type': request.POST.get("contract_type") or None,
            'status': request.POST.get("status") or None,
        }

        # تحويل الإنتاجية اليومية من نص لرقم لو موجود
        if data['production_capacity']:
            try:
                data['production_capacity'] = int(data['production_capacity'])
            except ValueError:
                data['production_capacity'] = None

        Makhbaz.objects.create(**data)
        messages.success(request, "✅ تم إضافة المخبز بنجاح")
        return redirect("makhabez_list")

    return render(request, "add_new_makhbaz.html")


@login_required(login_url='login')
def add_new_takiya(request):
    if request.method == "POST":
        name = request.POST.get("name")
        governorate = request.POST.get("governorate")
        address = request.POST.get("address")
        owner_name = request.POST.get("owner_name")
        owner_id = request.POST.get("owner_id")
        mobile_number = request.POST.get("mobile_number")
        coordinates = request.POST.get("coordinates")

        # تحويل القيم الرقمية إلى Integers مع معالجة الحقول الفارغة
        total_pots = int(request.POST.get("total_pots") or 0)
        pots_80 = int(request.POST.get("pots_80") or 0)
        pots_100 = int(request.POST.get("pots_100") or 0)
        pots_120 = int(request.POST.get("pots_120") or 0)
        pots_150 = int(request.POST.get("pots_150") or 0)
        pots_200 = int(request.POST.get("pots_200") or 0)
        daily_capacity = int(request.POST.get("daily_capacity") or 0)

        status = request.POST.get("status")

        # إنشاء التكية الجديدة
        Takiya.objects.create(
            name=name,
            governorate=governorate,
            address=address,
            owner_name=owner_name,
            owner_id=owner_id,
            mobile_number=mobile_number,
            coordinates=coordinates,
            total_pots=total_pots,
            pots_80=pots_80,
            pots_100=pots_100,
            pots_120=pots_120,
            pots_150=pots_150,
            pots_200=pots_200,
            daily_capacity=daily_capacity,
            status=status,
        )

        messages.success(request, "✅ تم إضافة التكية بنجاح.")
        return redirect("takiyat_list")

    return render(request, "add_new_takiya.html")


@login_required(login_url='login')
def delete_makhbaz(request, makhbaz_id):
    makhbaz = get_object_or_404(Makhbaz, id=makhbaz_id)

    # ⚠️ حذف المخبز
    makhbaz.delete()
    messages.success(request, "🗑️ تم حذف المخبز بنجاح.")
    return redirect("makhabez_list")



@login_required(login_url='login')
def delete_takiya(request, takiya_id):
    takiya = get_object_or_404(Takiya, id=takiya_id)

    # ⚠️ حذف التكية
    takiya.delete()
    messages.success(request, "🗑️ تم حذف التكية بنجاح.")
    return redirect("takiyat_list")  # اسم صفحة قائمة التكيات


logger = logging.getLogger(__name__)

@login_required(login_url='login')
def export_makhbaz_excel(request, makhbaz_id):
    try:
        makhbaz = get_object_or_404(Makhbaz, id=makhbaz_id)
        tasleemat = Taslima_makhbaz.objects.filter(makhbaz=makhbaz)

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المخبز"
        
        # تفعيل الاتجاه من اليمين لليسار (RTL)
        ws.rightToLeft = True
        ws.sheet_view.rightToLeft = True

        # تنسيقات الألوان المحسّنة
        TITLE_FILL = PatternFill(start_color='0B5394', end_color='0B5394', fill_type='solid')
        HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        SECTION_FILL = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
        LABEL_FILL = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
        DATA_FILL = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
        WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        TOTAL_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')

        # حدود محسّنة
        BORDER_THIN = Border(
            left=Side(style='thin', color='666666'),
            right=Side(style='thin', color='666666'),
            top=Side(style='thin', color='666666'),
            bottom=Side(style='thin', color='666666')
        )
        
        BORDER_THICK = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # خطوط عربية محسّنة
        title_font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
        section_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        label_font = Font(name='Arial', size=11, bold=True, color='1F4E78')
        normal_font = Font(name='Arial', size=10, bold=False, color='000000')
        bold_font = Font(name='Arial', size=10, bold=True, color='000000')
        total_font = Font(name='Arial', size=11, bold=True, color='990000')
        
        # ====================================================
        # العنوان الرئيسي
        # ====================================================
        ws.merge_cells('A1:J2')
        title_cell = ws['A1']
        title_cell.value = f"🥖  تقرير شامل لمخبز {makhbaz.name or 'غير محدد'}  🥖"
        title_cell.fill = TITLE_FILL
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = BORDER_THICK
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 35

        # ====================================================
        # قسم البيانات الأساسية
        # ====================================================
        start_row = 4
        ws.merge_cells(f'A{start_row}:J{start_row}')
        section_cell = ws[f'A{start_row}']
        section_cell.value = "📋  البيانات الأساسية للمخبز"
        section_cell.fill = SECTION_FILL
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        section_cell.border = BORDER_THICK
        ws.row_dimensions[start_row].height = 25

        # البيانات
        info_data = [
            ('اسم المخبز', makhbaz.name or "غير محدد", 'اسم صاحب المخبز', makhbaz.owner_name or "غير محدد"),
            ('رقم الهوية', makhbaz.owner_id or "غير محدد", 'رقم الجوال', makhbaz.mobile_number or "غير محدد"),
            ('العنوان', makhbaz.address or "غير محدد", 'المحافظة', makhbaz.governorate or "غير محدد"),
            ('الإحداثيات', makhbaz.coordinates or "غير محدد", 'نوع الفرن', makhbaz.oven_type or "غير محدد"),
            ('القدرة الإنتاجية', makhbaz.production_capacity or "غير محدد", 'نوع التعاقد', makhbaz.contract_type or "غير محدد"),
            ('حالة المخبز', makhbaz.status or "غير محدد", 'تاريخ الإضافة', makhbaz.created_at.strftime("%Y-%m-%d") if makhbaz.created_at else "غير محدد")
        ]

        current_row = start_row + 1
        for label1, value1, label2, value2 in info_data:
            # العمود الأول - التسمية
            cell = ws.cell(row=current_row, column=1)
            cell.value = label1
            cell.fill = LABEL_FILL
            cell.font = label_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # العمود الثاني - القيمة
            ws.merge_cells(f'B{current_row}:D{current_row}')
            cell = ws.cell(row=current_row, column=2)
            cell.value = value1
            cell.fill = WHITE_FILL
            cell.font = normal_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            for col in [3, 4]:
                ws.cell(row=current_row, column=col).border = BORDER_THIN
            
            # العمود الخامس - فاصل
            cell = ws.cell(row=current_row, column=5)
            cell.fill = SECTION_FILL
            cell.border = BORDER_THIN
            
            # العمود السادس - التسمية
            cell = ws.cell(row=current_row, column=6)
            cell.value = label2
            cell.fill = LABEL_FILL
            cell.font = label_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # الأعمدة 7-10 - القيمة
            ws.merge_cells(f'G{current_row}:J{current_row}')
            cell = ws.cell(row=current_row, column=7)
            cell.value = value2
            cell.fill = WHITE_FILL
            cell.font = normal_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            for col in [8, 9, 10]:
                ws.cell(row=current_row, column=col).border = BORDER_THIN
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # ====================================================
        # قسم التسليمات
        # ====================================================
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = "📦  سجل التسليمات والمواد المستلمة"
        section_cell.fill = SECTION_FILL
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        section_cell.border = BORDER_THICK
        ws.row_dimensions[current_row].height = 25

        # عناوين الأعمدة
        headers = ["التاريخ", "طحين\n(كغ)", "خميرة\n(كغ)", "ملح\n(كغ)", 
                   "سكر\n(كغ)", "زيت\n(لتر)", "حطب\n(كغ)", "غاز\n(كغ)", "إضافات"]
        
        header_row = current_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = header_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[header_row].height = 35

        # البيانات
        data_row = header_row + 1
        totals = {'flour': 0, 'yeast': 0, 'salt': 0, 'sugar': 0, 
                  'oil': 0, 'wood': 0, 'gas': 0}
        
        for idx, t in enumerate(tasleemat):
            flour = t.flour or 0
            yeast = t.yeast or 0
            salt = t.salt or 0
            sugar = t.sugar or 0
            oil = t.cooking_oil or 0
            wood = t.wood or 0
            gas = t.gas or 0
            
            row_data = [
                t.taslima_date.strftime("%Y-%m-%d") if t.taslima_date else "غير محدد",
                flour, yeast, salt, sugar, oil, wood, gas,
                t.additions or ""
            ]
            
            fill = WHITE_FILL if idx % 2 == 0 else DATA_FILL
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.value = value
                cell.border = BORDER_THIN
                cell.font = normal_font
                cell.fill = fill
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx in [2, 3, 4, 5, 6, 7, 8]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            totals['flour'] += flour
            totals['yeast'] += yeast
            totals['salt'] += salt
            totals['sugar'] += sugar
            totals['oil'] += oil
            totals['wood'] += wood
            totals['gas'] += gas
            
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # صف الإجمالي
        if tasleemat:
            total_row = data_row
            
            cell = ws.cell(row=total_row, column=1)
            cell.value = "📊 الإجمالي الكلي"
            cell.fill = TOTAL_FILL
            cell.font = total_font
            cell.border = BORDER_THICK
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            total_values = [totals['flour'], totals['yeast'], totals['salt'], 
                           totals['sugar'], totals['oil'], totals['wood'], 
                           totals['gas'], ""]
            
            for col_idx, total in enumerate(total_values, start=2):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value = total
                cell.fill = TOTAL_FILL
                cell.font = total_font
                cell.border = BORDER_THICK
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx < 9:
                    cell.number_format = '#,##0.00'
            
            ws.row_dimensions[total_row].height = 22

        # ضبط عرض الأعمدة
        column_widths = {
            'A': 16, 'B': 14, 'C': 14, 'D': 12, 'E': 12,
            'F': 14, 'G': 12, 'H': 12, 'I': 25, 'J': 2
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # تجميد الصفوف العلوية
        ws.freeze_panes = f'A{header_row + 1}'

        # إعداد الاستجابة
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        
        filename = f"تقرير_مخبز_{makhbaz.name or 'غير_محدد'}.xlsx"
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
        response['Content-Encoding'] = 'utf-8'

        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Excel for makhbaz {makhbaz_id}: {str(e)}")
        print(f"DEBUG - Error: {str(e)}")
        from django.contrib import messages
        messages.error(request, f"حدث خطأ أثناء تصدير الملف: {str(e)}")
        from django.shortcuts import redirect
        return redirect('makhabez_list')



logger = logging.getLogger(__name__)

@login_required(login_url='login')
def export_takiya_excel(request, takiya_id):
    try:
        takiya = get_object_or_404(Takiya, id=takiya_id)
        tasleemat = Taslima_takiya.objects.filter(takiya=takiya)

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير التكية"
        
        # تفعيل الاتجاه من اليمين لليسار (RTL)
        ws.rightToLeft = True
        ws.sheet_view.rightToLeft = True

        # تنسيقات الألوان المحسّنة
        TITLE_FILL = PatternFill(start_color='0B5394', end_color='0B5394', fill_type='solid')
        HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        SECTION_FILL = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
        LABEL_FILL = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
        DATA_FILL = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
        WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        TOTAL_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
        POT_FILL = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

        # حدود محسّنة
        BORDER_THIN = Border(
            left=Side(style='thin', color='666666'),
            right=Side(style='thin', color='666666'),
            top=Side(style='thin', color='666666'),
            bottom=Side(style='thin', color='666666')
        )
        
        BORDER_THICK = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # خطوط عربية محسّنة
        title_font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
        section_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        label_font = Font(name='Arial', size=11, bold=True, color='1F4E78')
        normal_font = Font(name='Arial', size=10, bold=False, color='000000')
        bold_font = Font(name='Arial', size=10, bold=True, color='000000')
        total_font = Font(name='Arial', size=11, bold=True, color='990000')
        pot_font = Font(name='Arial', size=10, bold=True, color='274E13')
        
        # ====================================================
        # العنوان الرئيسي
        # ====================================================
        ws.merge_cells('A1:P2')
        title_cell = ws['A1']
        title_cell.value = f"🍲  تقرير شامل لتكية {takiya.name or 'غير محدد'}  🍲"
        title_cell.fill = TITLE_FILL
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = BORDER_THICK
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 35

        # ====================================================
        # قسم البيانات الأساسية
        # ====================================================
        start_row = 4
        ws.merge_cells(f'A{start_row}:P{start_row}')
        section_cell = ws[f'A{start_row}']
        section_cell.value = "📋  البيانات الأساسية للتكية"
        section_cell.fill = SECTION_FILL
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        section_cell.border = BORDER_THICK
        ws.row_dimensions[start_row].height = 25

        # البيانات
        info_data = [
            ('اسم التكية', takiya.name or "غير محدد", 'اسم صاحب التكية', takiya.owner_name or "غير محدد"),
            ('رقم الهوية', takiya.owner_id or "غير محدد", 'رقم الجوال', takiya.mobile_number or "غير محدد"),
            ('العنوان', takiya.address or "غير محدد", 'المحافظة', takiya.governorate or "غير محدد"),
            ('الإحداثيات', takiya.coordinates or "غير محدد", 'القدرة الإنتاجية اليومية', f"{takiya.daily_capacity or 0} قدر" if takiya.daily_capacity else "غير محدد"),
            ('حالة التكية', takiya.status or "غير محدد", 'تاريخ الإضافة', takiya.created_at.strftime("%Y-%m-%d") if takiya.created_at else "غير محدد")
        ]

        current_row = start_row + 1
        for label1, value1, label2, value2 in info_data:
            # العمود الأول - التسمية
            cell = ws.cell(row=current_row, column=1)
            cell.value = label1
            cell.fill = LABEL_FILL
            cell.font = label_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # الأعمدة 2-6 - القيمة
            ws.merge_cells(f'B{current_row}:F{current_row}')
            cell = ws.cell(row=current_row, column=2)
            cell.value = value1
            cell.fill = WHITE_FILL
            cell.font = normal_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            for col in range(3, 7):
                ws.cell(row=current_row, column=col).border = BORDER_THIN
            
            # العمود السابع - فاصل
            cell = ws.cell(row=current_row, column=7)
            cell.fill = SECTION_FILL
            cell.border = BORDER_THIN
            
            # العمود الثامن - التسمية
            cell = ws.cell(row=current_row, column=8)
            cell.value = label2
            cell.fill = LABEL_FILL
            cell.font = label_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # الأعمدة 9-16 - القيمة
            ws.merge_cells(f'I{current_row}:P{current_row}')
            cell = ws.cell(row=current_row, column=9)
            cell.value = value2
            cell.fill = WHITE_FILL
            cell.font = normal_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            for col in range(10, 17):
                ws.cell(row=current_row, column=col).border = BORDER_THIN
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # ====================================================
        # قسم القدور والإمكانيات
        # ====================================================
        current_row += 1
        ws.merge_cells(f'A{current_row}:P{current_row}')
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = "🍯  القدور والإمكانيات"
        section_cell.fill = SECTION_FILL
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        section_cell.border = BORDER_THICK
        ws.row_dimensions[current_row].height = 25

        # عناوين القدور
        pot_headers = [
            'إجمالي القدور', 'قدور 80 لتر', 'قدور 100 لتر',
            'قدور 120 لتر', 'قدور 150 لتر', 'قدور 200 لتر'
        ]
        
        pot_values = [
            takiya.total_pots or 0,
            takiya.pots_80 or 0,
            takiya.pots_100 or 0,
            takiya.pots_120 or 0,
            takiya.pots_150 or 0,
            takiya.pots_200 or 0
        ]

        pot_row = current_row + 1
        for col_idx, header in enumerate(pot_headers, start=1):
            cell = ws.cell(row=pot_row, column=col_idx)
            cell.value = header
            cell.fill = LABEL_FILL
            cell.font = label_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[pot_row].height = 20

        pot_data_row = pot_row + 1
        for col_idx, value in enumerate(pot_values, start=1):
            cell = ws.cell(row=pot_data_row, column=col_idx)
            cell.value = value
            cell.fill = POT_FILL
            cell.font = pot_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0'
        
        ws.row_dimensions[pot_data_row].height = 20

        # ====================================================
        # قسم التسليمات
        # ====================================================
        current_row = pot_data_row + 2
        ws.merge_cells(f'A{current_row}:P{current_row}')
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = "📦  سجل التسليمات والمواد المستلمة"
        section_cell.fill = SECTION_FILL
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal='center', vertical='center')
        section_cell.border = BORDER_THICK
        ws.row_dimensions[current_row].height = 25

        # عناوين الأعمدة
        headers = [
            "التاريخ", "ملح\n(كغ)", "معكرونة\n(كغ)", "رز\n(كغ)",
            "زيت\n(لتر)", "بازيلا\n(علبة)", "عدس\n(كغ)", "لوبيا\n(كغ)",
            "صلصة\n(علبة)", "لانشون\n(علبة)", "بهار ماجي\n(كغ)", "شوربة خضار\n(كغ)",
            "٧ بهارات\n(كغ)", "سمنة\n(علبة)", "برغل\n(كغ)", "إضافات"
        ]
        
        header_row = current_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = header_font
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[header_row].height = 35

        # البيانات
        data_row = header_row + 1
        totals = {
            'salt': 0, 'macaroni': 0, 'rice': 0, 'oil': 0, 'peas': 0,
            'lentils': 0, 'beans': 0, 'sauce': 0, 'luncheon': 0,
            'maggi_spice': 0, 'vegetable_soup': 0, 'seven_spices': 0,
            'ghee': 0, 'bulgur': 0
        }
        
        for idx, t in enumerate(tasleemat):
            row_data = [
                t.taslima_date.strftime("%Y-%m-%d") if t.taslima_date else "غير محدد",
                t.salt or 0,
                t.macaroni or 0,
                t.rice or 0,
                t.oil or 0,
                t.peas or 0,
                t.lentils or 0,
                t.beans or 0,
                t.sauce or 0,
                t.luncheon or 0,
                t.maggi_spice or 0,
                t.vegetable_soup or 0,
                t.seven_spices or 0,
                t.ghee or 0,
                t.bulgur or 0,
                t.additions or ""
            ]
            
            # تحديث المجاميع
            totals['salt'] += t.salt or 0
            totals['macaroni'] += t.macaroni or 0
            totals['rice'] += t.rice or 0
            totals['oil'] += t.oil or 0
            totals['peas'] += t.peas or 0
            totals['lentils'] += t.lentils or 0
            totals['beans'] += t.beans or 0
            totals['sauce'] += t.sauce or 0
            totals['luncheon'] += t.luncheon or 0
            totals['maggi_spice'] += t.maggi_spice or 0
            totals['vegetable_soup'] += t.vegetable_soup or 0
            totals['seven_spices'] += t.seven_spices or 0
            totals['ghee'] += t.ghee or 0
            totals['bulgur'] += t.bulgur or 0
            
            fill = WHITE_FILL if idx % 2 == 0 else DATA_FILL
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.value = value
                cell.border = BORDER_THIN
                cell.font = normal_font
                cell.fill = fill
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx in range(2, 16):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # صف الإجمالي
        if tasleemat:
            total_row = data_row
            
            cell = ws.cell(row=total_row, column=1)
            cell.value = "📊 الإجمالي الكلي"
            cell.fill = TOTAL_FILL
            cell.font = total_font
            cell.border = BORDER_THICK
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            total_values = [
                totals['salt'], totals['macaroni'], totals['rice'], totals['oil'],
                totals['peas'], totals['lentils'], totals['beans'], totals['sauce'],
                totals['luncheon'], totals['maggi_spice'], totals['vegetable_soup'],
                totals['seven_spices'], totals['ghee'], totals['bulgur'], ""
            ]
            
            for col_idx, total in enumerate(total_values, start=2):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value = total
                cell.fill = TOTAL_FILL
                cell.font = total_font
                cell.border = BORDER_THICK
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx < 16:
                    cell.number_format = '#,##0'
            
            ws.row_dimensions[total_row].height = 22

        # ضبط عرض الأعمدة
        column_widths = {
            'A': 14, 'B': 10, 'C': 11, 'D': 10, 'E': 11,
            'F': 11, 'G': 10, 'H': 10, 'I': 11, 'J': 11,
            'K': 11, 'L': 12, 'M': 11, 'N': 11, 'O': 10, 'P': 20
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # تجميد الصفوف العلوية
        ws.freeze_panes = f'A{header_row + 1}'

        # إعداد الاستجابة
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        
        filename = f"تقرير_تكية_{takiya.name or 'غير_محدد'}.xlsx"
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
        response['Content-Encoding'] = 'utf-8'

        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Excel for takiya {takiya_id}: {str(e)}")
        print(f"DEBUG - Error: {str(e)}")
        from django.contrib import messages
        messages.error(request, f"حدث خطأ أثناء تصدير الملف: {str(e)}")
        from django.shortcuts import redirect
        return redirect('takaya_list')
